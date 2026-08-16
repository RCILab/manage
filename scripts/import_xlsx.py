#!/usr/bin/env python3
"""
「연구비 소요 내역 (YYYY)」 구글 시트를 XLSX 로 내려받아 Supabase 초기 데이터(SQL) 로 변환한다.

사용법:
    pip install openpyxl
    python scripts/import_xlsx.py "연구비 소요 내역 (2026).xlsx" --year 2026 --out local/seed.sql

생성된 local/seed.sql 을 Supabase SQL Editor 에서 실행한다 (schema.sql 을 먼저 실행할 것).
※ seed.sql 에는 개인 급여 정보가 들어 있으므로 절대 git 에 커밋하지 말 것 (local/ 은 .gitignore 됨).

시트 구조 (2026 기준):
  - '2026 인건비 '        : 학생 × 과제 × 월 배분 그리드 (+ '기준' 열)
  - '균등차등 인건비'      : 비고(신분) 참고용
  - 과제 탭 (예: '아바타(김상현)') : 상단 연차별 계획(천원), 비목별 계획/이월(원), 하단 원장
"""
import argparse
import datetime as dt
import re
import sys

import openpyxl

# ---------------------------------------------------------------------------
# 과제 매핑
# ---------------------------------------------------------------------------
# 인건비 탭 헤더(줄바꿈 앞 부분) → 과제 코드
HEADER_TO_CODE = {
    '기본연구': 'basic', '초격차': 'cgc', '기초연구실': 'brl', '한수원': 'khnp', 'DCP': 'dcp',
    '로봇테스트필드': 'testfield', '글로벌TOP': 'gtop', '로봇핸드': 'hand', '아바타': 'avatar',
    'KIST': 'kist', '신진': 'sinjin', '농수': 'nongsu', '인력양성': 'hrd', 'RISE': 'rise',
    '이월금': 'carryover', 'KRISO': 'kriso', '스케일업': 'scaleup', 'BK장학금': 'bk',
    'RA, TA외 장학금': 'rata', 'SW': 'sw', '티보로틱스': 'tibo', '한중': 'hanjung',
}
# 코드 → (표시 이름, 정렬)
PROJECT_NAMES = {
    'basic': '기본연구', 'cgc': '초격차', 'brl': '기초연구실', 'khnp': '한수원', 'dcp': 'DCP',
    'testfield': '로봇테스트필드', 'gtop': '글로벌TOP', 'hand': '로봇핸드', 'avatar': '아바타',
    'kist': 'KIST', 'sinjin': '신진', 'nongsu': '농수', 'hrd': '인력양성', 'rise': 'RISE',
    'carryover': '이월금', 'kriso': 'KRISO', 'scaleup': '스케일업', 'bk': 'BK장학금',
    'rata': 'RA/TA 장학금', 'sw': 'SW', 'tibo': '티보로틱스', 'hanjung': '한중',
}
# 과제 탭 → (코드, 현재 연차 여부, (기간시작, 기간끝) 또는 None, 과제 종료 여부)
# 기간을 모르는 현재 연차는 --year 달력연도, 과거 연차는 (year-1) 달력연도로 채운다.
TABS = [
    ('아바타(김상현)',              'avatar',    True,  None,                         False),
    ('기초연구실(남창주)',          'brl',       True,  None,                         False),
    ('초격차(김진균)',              'cgc',       True,  None,                         False),
    ('기초연구실(남창주, 2차년도)', 'brl',       False, 'prev',                       False),
    ('한수원(김상현)',              'khnp',      True,  None,                         False),
    ('DCP(이승훈)',                 'dcp',       True,  None,                         False),
    ('한수원(김상현, 2차년도)',     'khnp',      False, 'prev',                       False),
    ('테스트필드(임성수)',          'testfield', True,  None,                         False),
    ('로봇핸드(김상현)',            'hand',      True,  ('2026-06-01', '2027-03-31'), False),
    ('로봇핸드(김상현, 1차년도)',   'hand',      False, ('2025-09-01', '2026-05-31'), False),
    ('글로벌TOP(김상현)',           'gtop',      True,  ('2026-05-01', '2027-04-30'), False),
    ('KIST(김상현)',                'kist',      True,  None,                         False),
    ('글로벌TOP(김상현, 1차년도)',  'gtop',      False, ('2025-05-01', '2026-04-30'), False),
    ('KRISO(김상현)',               'kriso',     True,  ('2026-06-01', '2026-11-30'), False),
    ('농수(김상현)',                'nongsu',    True,  None,                         False),
    ('티보로틱스(김상현)',          'tibo',      True,  None,                         False),
    ('신진(김상현)',                'sinjin',    True,  None,                         False),
    ('RISE(김상현)',                'rise',      True,  ('2026-06-01', '2026-11-30'), False),
    ('인력양성(임성수)',            'hrd',       True,  None,                         False),
    ('스케일업(김상현)',            'scaleup',   True,  ('2026-07-01', '2026-12-31'), False),
    ('한중(김상현)',                'hanjung',   True,  ('2026-10-01', '2026-12-31'), False),
    ('BK(김영선)',                  'bk',        True,  None,                         False),
    ('기본연구(김상현, 종료)',      'basic',     True,  None,                         True),
]
# 탭이 없는 다른 연차의 기간 (알려진 것만)
EXTRA_YEAR_PERIODS = {
    ('rise', 1): ('2025-11-01', '2026-01-31'),
}

# 시트 비목 표의 소그룹(C열) → 하위 항목이 없을 때 쓸 비목명
GROUP_LEAF_FALLBACK = {'총합': None}


# ---------------------------------------------------------------------------
# 유틸
# ---------------------------------------------------------------------------
def q(s):
    """SQL 문자열 리터럴"""
    if s is None:
        return 'null'
    return "'" + str(s).replace("'", "''") + "'"


def num(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(round(v))
    if isinstance(v, str):
        s = v.replace(',', '').replace('₩', '').strip()
        if s in ('', '-'):
            return None
        try:
            return int(round(float(s)))
        except ValueError:
            return None
    return None


def txt(v):
    if v is None:
        return None
    s = str(v).replace('\n', ' ').strip()
    return s or None


def parse_date(v):
    """'26.03.11' / '2026.08.03' / '26.7.29' / datetime → (date_str or None, 원문)"""
    if v is None:
        return None, None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime('%Y-%m-%d'), None
    s = str(v).strip()
    m = re.match(r'^(\d{2,4})[.\-/](\d{1,2})[.\-/](\d{1,2})\.?$', s)
    if not m:
        return None, s
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    if not (1900 < y < 2100 and 1 <= mo <= 12 and 1 <= d <= 31):
        return None, s
    try:
        return dt.date(y, mo, d).strftime('%Y-%m-%d'), None
    except ValueError:
        return None, s


def year_no_of(v):
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str):
        m = re.match(r'^\s*(\d+)', v)
        if m:
            return int(m.group(1))
    return None


# ---------------------------------------------------------------------------
# 파서
# ---------------------------------------------------------------------------
def parse_labor(ws, year):
    """'YYYY 인건비 ' 탭 → members, allocations, targets"""
    header = {}
    for c in range(4, ws.max_column + 1):
        v = ws.cell(3, c).value
        if isinstance(v, str) and v.strip():
            key = v.split('\n')[0].strip()
            header[c] = key
    proj_cols = {}
    for c, key in header.items():
        if key in HEADER_TO_CODE:
            proj_cols[c] = HEADER_TO_CODE[key]
    # '기준' 열 = '기준대비 증감액' 헤더 열 (금액 행에는 기준 금액이 들어 있음)
    target_col = next((c for c, k in header.items() if k.startswith('기준대비')), None)

    members, allocs, targets = [], [], []
    r = 5
    order = 0
    while r <= ws.max_row:
        name_raw = ws.cell(r, 2).value
        if not (isinstance(name_raw, str) and name_raw.strip()):
            r += 1
            continue
        raw = name_raw.strip()
        is_bk = '(BK)' in raw
        name = re.sub(r'\(BK\)', '', raw).replace('\n', ' ').strip()
        order += 1
        members.append({'name': name, 'is_bk': is_bk, 'sort_order': order})
        for m in range(1, 13):
            row = r + 2 * (m - 1)
            for c, code in proj_cols.items():
                a = num(ws.cell(row, c).value)
                if a:
                    allocs.append((name, code, year, m, a))
            if target_col:
                t = num(ws.cell(row, target_col).value)
                if t:
                    targets.append((name, year, m, t))
        r += 27
    used_codes = sorted({a[1] for a in allocs})
    return members, allocs, targets, used_codes


def parse_notes(wb):
    """'균등차등 인건비' 탭의 비고(신분) → {name: note}"""
    notes = {}
    if '균등차등 인건비' not in wb.sheetnames:
        return notes
    ws = wb['균등차등 인건비']
    cur = None
    for r in range(1, ws.max_row + 1):
        n = ws.cell(r, 2).value
        if isinstance(n, str) and n.strip():
            cur = n.strip()
        v = ws.cell(r, 10).value
        if cur and isinstance(v, str) and v.strip() and v.strip() != '비고':
            notes.setdefault(cur, v.strip())
    return notes


def position_from_note(note, is_bk):
    if not note:
        return None
    if '학부' in note:
        return '학부연구생'
    if '박사' in note:
        return '박사'
    if '석박' in note:
        return '석박통합'
    if '석사' in note:
        return '석사'
    if '직장' in note:
        return '연구원'
    return None


def parse_project_tab(ws):
    """과제 탭 → dict(cur_year_no, contact, plans[], budgets[], ledger[], total_note)"""
    out = {'plans': [], 'budgets': [], 'ledger': [], 'contact': None, 'cur_year_no': None}
    out['cur_year_no'] = year_no_of(ws.cell(3, 15).value)
    # 연차별 계획 (천원): B=연차, C='현금', J=직접비소계, K=간접비, M=할당예산 총액
    for r in range(5, 22):
        if ws.cell(r, 3).value != '현금':
            continue
        yn = year_no_of(ws.cell(r, 2).value)
        if yn is None:
            continue
        direct = num(ws.cell(r, 10).value)
        indirect = num(ws.cell(r, 11).value)
        total = num(ws.cell(r, 13).value)
        if total is None:
            total = num(ws.cell(r, 12).value)
        if direct is None and indirect is None and total is None:
            continue
        out['plans'].append({
            'year_no': yn,
            'direct': direct * 1000 if direct is not None else None,
            'indirect': indirect * 1000 if indirect is not None else None,
            'total': total * 1000 if total is not None else None,
        })
    # 담당 선생님
    for r in range(20, 30):
        for c in range(4, 8):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith('담당'):
                m = re.search(r'\((.*?)\)', v)
                out['contact'] = m.group(1).strip() if m else v.replace('담당 선생님', '').strip()
    # 비목별 표: '비목' 헤더 행 찾기
    hdr_row = None
    for r in range(15, 40):
        if ws.cell(r, 2).value == '비목':
            hdr_row = r
            break
    if hdr_row:
        cols = {}
        for c in range(3, 10):
            v = ws.cell(hdr_row + 1, c).value
            if v in ('계획', '이월', '사용', '잔액'):
                cols[v] = c
        group = None
        r = hdr_row + 2
        while r < hdr_row + 60:
            b, c3, d = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value
            if isinstance(b, str) and b.startswith('총계'):
                break
            if isinstance(b, str) and b.strip():
                group = b.strip()
            leaf = txt(d) or txt(c3)
            if leaf and leaf not in ('총합',):
                planned = num(ws.cell(r, cols.get('계획', 5)).value)
                carry = num(ws.cell(r, cols.get('이월', 6)).value)
                if planned or carry:
                    out['budgets'].append({'category': leaf, 'parent': group, 'planned': planned or 0,
                                           'carryover': carry or 0, 'order': r})
            r += 1
        # 원장: '순서' 헤더
        led_row = None
        for r in range(hdr_row + 2, hdr_row + 80):
            if ws.cell(r, 2).value == '순서' and ws.cell(r, 3).value == '사용일자':
                led_row = r
                break
        if led_row:
            r = led_row + 1
            empty_streak = 0
            while r <= ws.max_row and empty_streak < 5:
                date_v = ws.cell(r, 3).value
                cat = txt(ws.cell(r, 4).value)
                amt = num(ws.cell(r, 5).value)
                memo = txt(ws.cell(r, 6).value)
                if amt is None and not cat and not memo and date_v in (None, ''):
                    empty_streak += 1
                    r += 1
                    continue
                empty_streak = 0
                if amt is not None and (cat or memo):
                    d, raw = parse_date(date_v)
                    if raw:
                        memo = f'[날짜확인: {raw}] ' + (memo or '')
                    out['ledger'].append({'date': d, 'category': cat, 'amount': amt, 'memo': memo})
                r += 1
    # 총액 메모 (T22/U22)
    for r in range(20, 24):
        for c in range(18, 24):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith('총액'):
                out['total_note'] = f"{v} {num(ws.cell(r, c + 1).value) or ''}".strip()
    return out


def parse_carryover(wb):
    if '이월금(김상현)' not in wb.sheetnames:
        return None
    ws = wb['이월금(김상현)']
    vals = {}
    for r in range(1, 12):
        for c in range(1, 12):
            v = ws.cell(r, c).value
            if v in ('전년도 이월', '올해 이월 예상', '학교 반납'):
                vals[v] = num(ws.cell(r + 1, c).value)
    if not vals:
        return None
    return ' / '.join(f'{k} {v:,}' for k, v in vals.items() if v is not None)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx')
    ap.add_argument('--year', type=int, default=2026)
    ap.add_argument('--out', default='local/seed.sql')
    ap.add_argument('--labor-sheet', default=None, help="기본값: 'YYYY 인건비 '")
    args = ap.parse_args()
    year = args.year

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    labor_name = args.labor_sheet or next(
        (n for n in wb.sheetnames if n.strip() == f'{year} 인건비'), None)
    if not labor_name:
        sys.exit(f"인건비 탭을 찾을 수 없음: '{year} 인건비'")

    members, allocs, targets, used_codes = parse_labor(wb[labor_name], year)
    notes = parse_notes(wb)
    carry_note = parse_carryover(wb)

    # 과제 탭 파싱
    tabs = []
    for tab, code, is_cur, period, ended in TABS:
        if tab not in wb.sheetnames:
            print(f'  (탭 없음, 건너뜀) {tab}', file=sys.stderr)
            continue
        info = parse_project_tab(wb[tab])
        tabs.append((tab, code, is_cur, period, ended, info))

    # 과제 목록: 탭에 있는 것 + 인건비 배분에 쓰인 것 (0 만 있는 열은 제외)
    codes = []
    for _, code, *_ in tabs:
        if code not in codes:
            codes.append(code)
    for code in used_codes:
        if code not in codes:
            codes.append(code)
    if 'carryover' not in codes and any(a[1] == 'carryover' for a in allocs):
        codes.append('carryover')

    ended_codes = {code for _, code, _, _, ended, _ in tabs if ended}
    contacts = {}
    for _, code, is_cur, _, _, info in tabs:
        if info.get('contact') and (is_cur or code not in contacts):
            contacts[code] = info['contact']
    pi_of = {}
    for tab, code, *_ in tabs:
        m = re.search(r'\((.*?)[,)]', tab)
        if m:
            pi_of[code] = m.group(1).strip()
    pi_of.setdefault('carryover', '김상현')

    L = []
    w = L.append
    w('-- 자동 생성: import_xlsx.py  (원본: %s, 연도 %d)' % (args.xlsx, year))
    w('-- ※ 개인 급여 정보 포함. git 에 커밋 금지.')
    w('begin;')

    # ---- projects
    w('\n-- 과제')
    for i, code in enumerate(codes):
        name = PROJECT_NAMES.get(code, code)
        w(f"insert into public.projects (code, name, pi_name, agency_contact, active, sort_order) values "
          f"({q(code)}, {q(name)}, {q(pi_of.get(code))}, {q(contacts.get(code))}, "
          f"{'false' if code in ended_codes else 'true'}, {i * 10}) "
          f"on conflict (code) do update set name = excluded.name, pi_name = coalesce(excluded.pi_name, projects.pi_name), "
          f"agency_contact = coalesce(excluded.agency_contact, projects.agency_contact), sort_order = excluded.sort_order;")

    # ---- project_years
    w('\n-- 과제 연차')
    prev_period = (f'{year - 1}-01-01', f'{year - 1}-12-31')
    cur_period = (f'{year}-01-01', f'{year}-12-31')
    py_seen = set()

    def py_upsert(code, yn, label, period, is_cur, plan, note):
        ps, pe = (period or (None, None))
        pd = plan or {}
        w(f"insert into public.project_years (project_id, year_no, label, period_start, period_end, is_current, plan_direct, plan_indirect, plan_total, note) "
          f"select id, {yn}, {q(label)}, {q(ps)}, {q(pe)}, {'true' if is_cur else 'false'}, "
          f"{pd.get('direct') if pd.get('direct') is not None else 'null'}, "
          f"{pd.get('indirect') if pd.get('indirect') is not None else 'null'}, "
          f"{pd.get('total') if pd.get('total') is not None else 'null'}, {q(note)} "
          f"from public.projects where code = {q(code)} "
          f"on conflict (project_id, year_no) do update set label = excluded.label, "
          f"period_start = coalesce(excluded.period_start, project_years.period_start), "
          f"period_end = coalesce(excluded.period_end, project_years.period_end), "
          f"is_current = excluded.is_current or project_years.is_current, "
          f"plan_direct = coalesce(excluded.plan_direct, project_years.plan_direct), "
          f"plan_indirect = coalesce(excluded.plan_indirect, project_years.plan_indirect), "
          f"plan_total = coalesce(excluded.plan_total, project_years.plan_total), "
          f"note = coalesce(excluded.note, project_years.note);")

    for tab, code, is_cur, period, ended, info in tabs:
        yn = info['cur_year_no'] or 1
        if period == 'prev':
            period = prev_period
        elif period is None:
            period = cur_period
        plans = {p['year_no']: p for p in info['plans']}
        note = info.get('total_note')
        py_upsert(code, yn, f'{yn}차년도', period, is_cur, plans.get(yn), note)
        py_seen.add((code, yn))
        for oyn, p in plans.items():
            if (code, oyn) in py_seen:
                continue
            py_upsert(code, oyn, f'{oyn}차년도', EXTRA_YEAR_PERIODS.get((code, oyn)), False, p, None)
            py_seen.add((code, oyn))
    for code in codes:
        if not any(c == code for c, _ in py_seen):
            note = carry_note if code == 'carryover' else None
            py_upsert(code, 1, '1차년도', cur_period, True, None, note)
            py_seen.add((code, 1))

    # ---- categories (시트에만 있는 비목 추가)
    w('\n-- 비목')
    cat_seen = set()
    for _, _, _, _, _, info in tabs:
        for b in info['budgets']:
            if b['category'] not in cat_seen:
                cat_seen.add(b['category'])
                w(f"insert into public.categories (name, parent, sort_order) values ({q(b['category'])}, {q(b['parent'])}, {b['order'] * 10}) on conflict (name) do nothing;")
    for _, _, _, _, _, info in tabs:
        for l in info['ledger']:
            if l['category'] and l['category'] not in cat_seen:
                cat_seen.add(l['category'])
                w(f"insert into public.categories (name, parent, sort_order) values ({q(l['category'])}, null, 900) on conflict (name) do nothing;")

    # ---- budget_lines & ledger
    w('\n-- 비목별 예산 / 원장')
    for tab, code, is_cur, period, ended, info in tabs:
        yn = info['cur_year_no'] or 1
        sel_py = (f"(select py.id from public.project_years py join public.projects p on p.id = py.project_id "
                  f"where p.code = {q(code)} and py.year_no = {yn})")
        for b in info['budgets']:
            w(f"insert into public.budget_lines (project_year_id, category, planned, carryover) "
              f"select {sel_py}, {q(b['category'])}, {b['planned']}, {b['carryover']} "
              f"on conflict (project_year_id, category) do update set planned = excluded.planned, carryover = excluded.carryover;")
        if info['ledger']:
            w(f"delete from public.ledger where project_year_id = {sel_py} and created_by = 'import';")
            for l in info['ledger']:
                w(f"insert into public.ledger (project_year_id, spent_on, category, amount, memo, created_by) "
                  f"values ({sel_py}, {q(l['date'])}, {q(l['category'])}, {l['amount']}, {q(l['memo'])}, 'import');")

    # ---- members
    w('\n-- 구성원 (이메일은 관리자 화면에서 입력)')
    for m in members:
        note = notes.get(m['name'])
        pos = position_from_note(note, m['is_bk'])
        w(f"insert into public.members (name, role, position, is_bk, note, sort_order) "
          f"select {q(m['name'])}, 'student', {q(pos)}, {'true' if m['is_bk'] else 'false'}, {q(note)}, {m['sort_order']} "
          f"where not exists (select 1 from public.members where name = {q(m['name'])});")
    w("insert into public.members (name, role, position, sort_order) select '행정조교', 'admin', '행정', -50 "
      "where not exists (select 1 from public.members where role = 'admin');")

    # ---- allocations
    w('\n-- 학생 인건비 배분')
    for name, code, y, mo, amt in allocs:
        w(f"insert into public.allocations (member_id, project_id, year, month, amount) "
          f"select m.id, p.id, {y}, {mo}, {amt} from public.members m, public.projects p "
          f"where m.name = {q(name)} and p.code = {q(code)} "
          f"on conflict (member_id, project_id, year, month) do update set amount = excluded.amount;")

    # ---- targets
    w('\n-- 학생별 월 기준 금액')
    for name, y, mo, amt in targets:
        w(f"insert into public.salary_targets (member_id, year, month, amount) "
          f"select id, {y}, {mo}, {amt} from public.members where name = {q(name)} "
          f"on conflict (member_id, year, month) do update set amount = excluded.amount;")

    w('\ncommit;')

    import os
    os.makedirs(os.path.dirname(args.out) or '.', exist_ok=True)
    with open(args.out, 'w', encoding='utf-8') as f:
        f.write('\n'.join(L) + '\n')

    n_led = sum(len(i['ledger']) for *_, i in tabs)
    n_bud = sum(len(i['budgets']) for *_, i in tabs)
    print(f'완료: {args.out}')
    print(f'  과제 {len(codes)}개, 연차 {len(py_seen)}개, 비목예산 {n_bud}행, 원장 {n_led}행, '
          f'학생 {len(members)}명, 배분 {len(allocs)}셀, 기준 {len(targets)}셀')


if __name__ == '__main__':
    main()
